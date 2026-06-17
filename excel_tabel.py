from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
import calendar
from database import xodimlar_royxati, xodim_oylik_tabel, oylik_maosh_olish, oylik_maosh_hisoblash

QIZIL = "C00000"
YASHIL = "375623"
SARIQ = "FFC000"
KOK = "1F3864"
KULRANG = "D9D9D9"
OCHIQ_YASHIL = "E2EFDA"
OCHIQ_QIZIL = "FCE4D6"
OCHIQ_KOK = "DEEAF1"

def chegara(qalinlik='thin'):
    s = Side(style=qalinlik)
    return Border(left=s, right=s, top=s, bottom=s)

def sarlavha_uslub(cell, rang, matn_rangi="FFFFFF", qalin=True, markaziy=True):
    cell.fill = PatternFill("solid", start_color=rang, end_color=rang)
    cell.font = Font(bold=qalin, color=matn_rangi, name="Arial", size=10)
    cell.border = chegara()
    if markaziy:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def holat_rang(holat):
    ranglar = {
        'Keldi': OCHIQ_YASHIL,
        'Kelmadi': OCHIQ_QIZIL,
        'Kasallik': OCHIQ_KOK,
        "Ta'til": "FFF2CC",
        'Dam olish': KULRANG,
    }
    return ranglar.get(holat, "FFFFFF")

def oylik_tabel_excel(oy):
    """Oylik tabel - barcha xodimlar"""
    yil, oy_raqam = map(int, oy.split('-'))
    kunlar_soni = calendar.monthrange(yil, oy_raqam)[1]
    oy_nomi = datetime(yil, oy_raqam, 1).strftime('%B %Y')
    
    wb = Workbook()
    
    # ===== 1-VARAQ: TABEL =====
    ws = wb.active
    ws.title = "Tabel"
    
    # Sarlavha
    ws.merge_cells(f'A1:{get_column_letter(kunlar_soni*2+3)}1')
    ws['A1'] = f"ISHCHI KUCHI TABELI — {oy_nomi.upper()}"
    ws['A1'].font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    ws['A1'].fill = PatternFill("solid", start_color=KOK, end_color=KOK)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Ustun sarlavhalari
    ws['A2'] = '№'
    ws['B2'] = 'Xodim F.I.O'
    ws['C2'] = 'Lavozim'
    
    col = 4
    for kun in range(1, kunlar_soni + 1):
        d = date(yil, oy_raqam, kun)
        hafta = d.strftime('%a')
        ws.merge_cells(f'{get_column_letter(col)}2:{get_column_letter(col+1)}2')
        cell = ws.cell(row=2, column=col, value=f'{kun}\n{hafta}')
        rang = OCHIQ_QIZIL if d.weekday() >= 5 else KULRANG
        sarlavha_uslub(cell, rang, matn_rangi="000000")
        ws.column_dimensions[get_column_letter(col)].width = 4
        ws.column_dimensions[get_column_letter(col+1)].width = 4
        col += 2
    
    # Jami ustunlar
    jami_col = col
    for sarlavha in ['Jami\nKun', 'Jami\nSoat', 'Kech\nDaq', 'Kasallik', "Ta'til", 'Kelmadi']:
        c = ws.cell(row=2, column=jami_col, value=sarlavha)
        sarlavha_uslub(c, KOK)
        ws.column_dimensions[get_column_letter(jami_col)].width = 8
        jami_col += 1
    
    # Ustun kengliklarini sozlash
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.row_dimensions[2].height = 35
    ws.freeze_panes = 'D3'
    
    # Xodimlar ma'lumotlari
    xodimlar = xodimlar_royxati()
    qator = 3
    
    for i, xodim in enumerate(xodimlar, 1):
        xid, ism, lavozim = xodim[0], xodim[1], xodim[2]
        
        ws.cell(row=qator, column=1, value=i).border = chegara()
        ws.cell(row=qator, column=2, value=ism).border = chegara()
        ws.cell(row=qator, column=3, value=lavozim).border = chegara()
        
        # Tabel ma'lumotlari
        tabel = xodim_oylik_tabel(xid, oy)
        # {(sana, smen): holat}
        tabel_dict = {}
        kechikish_dict = {}
        for t in tabel:
            tabel_dict[(t[2], t[3])] = t[4]
            kechikish_dict[(t[2], t[3])] = t[6] or 0
        
        jami_kun = 0
        jami_kechikish = 0
        kasallik = 0
        tatil = 0
        kelmadi = 0
        
        col = 4
        for kun in range(1, kunlar_soni + 1):
            sana = f"{oy}-{kun:02d}"
            d = date(yil, oy_raqam, kun)
            
            # 2 smen
            for smen in ['1-smen', '2-smen']:
                holat = tabel_dict.get((sana, smen), '')
                kech = kechikish_dict.get((sana, smen), 0)
                
                cell = ws.cell(row=qator, column=col)
                if holat:
                    if holat == 'Keldi':
                        cell.value = 'K'
                        jami_kun += 0.5
                        jami_kechikish += kech
                    elif holat == 'Kelmadi':
                        cell.value = '—'
                        kelmadi += 0.5
                    elif holat == 'Kasallik':
                        cell.value = 'Kas'
                        kasallik += 0.5
                    elif holat == "Ta'til":
                        cell.value = 'T'
                        tatil += 0.5
                    cell.fill = PatternFill("solid", 
                        start_color=holat_rang(holat), 
                        end_color=holat_rang(holat))
                elif d.weekday() >= 5:
                    cell.fill = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
                
                cell.border = chegara()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial", size=8)
                col += 1
        
        # Jami ustunlar
        jami_soat = jami_kun * 12
        for val in [int(jami_kun), int(jami_soat), int(jami_kechikish), 
                    int(kasallik), int(tatil), int(kelmadi)]:
            c = ws.cell(row=qator, column=col, value=val)
            c.border = chegara()
            c.alignment = Alignment(horizontal="center")
            c.font = Font(name="Arial", size=10, bold=True)
            col += 1
        
        ws.row_dimensions[qator].height = 18
        qator += 1
    
    # ===== 2-VARAQ: MAOSH =====
    ws2 = wb.create_sheet("Maosh hisobi")
    
    ws2.merge_cells('A1:I1')
    ws2['A1'] = f"MAOSH HISOBI — {oy_nomi.upper()}"
    ws2['A1'].font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    ws2['A1'].fill = PatternFill("solid", start_color=KOK, end_color=KOK)
    ws2['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30
    
    sarlavhalar = ['№', 'Xodim F.I.O', 'Lavozim', 'Soatlik\n(so\'m)', 
                   'Jami soat', 'Kechikish\n(daq)', 'Kasallik\n(kun)',
                   "Ta'til\n(kun)", 'Hisoblangan\nmaosh (so\'m)']
    kengliklar = [4, 22, 14, 12, 10, 10, 10, 10, 18]
    
    for col, (s, k) in enumerate(zip(sarlavhalar, kengliklar), 1):
        c = ws2.cell(row=2, column=col, value=s)
        sarlavha_uslub(c, KOK)
        ws2.column_dimensions[get_column_letter(col)].width = k
    ws2.row_dimensions[2].height = 35
    ws2.freeze_panes = 'A3'
    
    # Maosh hisoblash
    oylik_maosh_hisoblash(oy)
    maoshlar = oylik_maosh_olish(oy)
    
    jami_maosh = 0
    for i, m in enumerate(maoshlar, 1):
        qator = i + 2
        fill = PatternFill("solid", start_color="F9F9F9" if i%2==0 else "FFFFFF",
                          end_color="F9F9F9" if i%2==0 else "FFFFFF")
        
        qiymatlar = [i, m[10], m[11], f"{m[12]:,}", 
                     f"{m[2]:.1f}", m[5], m[3], m[4], f"{m[7]:,}"]
        
        for col, val in enumerate(qiymatlar, 1):
            c = ws2.cell(row=qator, column=col, value=val)
            c.border = chegara()
            c.fill = fill
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="center" if col != 2 else "left",
                                   vertical="center")
        
        # Maosh ustunini ajratib ko'rsatish
        maosh_cell = ws2.cell(row=qator, column=9)
        maosh_cell.font = Font(name="Arial", size=10, bold=True, color=QIZIL)
        jami_maosh += m[7]
        ws2.row_dimensions[qator].height = 18
    
    # Jami qator
    jami_qator = len(maoshlar) + 3
    ws2.merge_cells(f'A{jami_qator}:H{jami_qator}')
    c = ws2.cell(row=jami_qator, column=1, value="JAMI MAOSH FONDI:")
    sarlavha_uslub(c, KOK)
    c = ws2.cell(row=jami_qator, column=9, value=f"{jami_maosh:,}")
    sarlavha_uslub(c, QIZIL)
    
    # ===== 3-VARAQ: IZOH =====
    ws3 = wb.create_sheet("Belgilar")
    ws3['A1'] = "BELGILAR IZOHI"
    ws3['A1'].font = Font(bold=True, size=12, name="Arial")
    
    belgilar = [
        ('K', 'Keldi (smen ishladi)', OCHIQ_YASHIL),
        ('—', 'Kelmadi (sababsiz)', OCHIQ_QIZIL),
        ('Kas', 'Kasallik', OCHIQ_KOK),
        ('T', "Ta'tilda", "FFF2CC"),
        ('', 'Dam olish kuni', KULRANG),
    ]
    for i, (belgi, izoh, rang) in enumerate(belgilar, 3):
        c1 = ws3.cell(row=i, column=1, value=belgi)
        c1.fill = PatternFill("solid", start_color=rang, end_color=rang)
        c1.border = chegara()
        c1.alignment = Alignment(horizontal="center")
        ws3.cell(row=i, column=2, value=izoh).font = Font(name="Arial", size=10)
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 25
    
    fayl = f"tabel_{oy}.xlsx"
    wb.save(fayl)
    return fayl
